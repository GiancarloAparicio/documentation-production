from docxtpl import DocxTemplate
from editpyxl import Workbook
from datetime import datetime
from openpyxl import load_workbook
import json
import pandas as pd
import os
import jinja2
import re


matchs = [
    {"name": "Ruta", "key": 270, "rule": "^Ruta.*"},
    {
        "name": "URL",
        "key": 271,
        "rule": "^URL.*",
    },
    {
        "name": "Email",
        "key": 272,
        "rule": "^Email.*",
    },
    {"name": "Flag", "key": 274, "rule": "^Flag.*"},
    {"name": "Numero", "key": 275, "rule": "^Numero.*"},
    {"name": "Asset", "key": 277, "rule": "^Asset.*"},
    {"name": "Usuario", "key": 278, "rule": "^Usuario.*"},
    {"name": "Servidor", "key": 279, "rule": "^Servidor.*"},
    {
        "name": "Nombre_arch",
        "key": 273,
        "rule": "^Nombre_arch.*",
    },
    {"name": "Store_Procedure", "key": 280, "rule": "^Store_Procedure.*"},
    {
        "name": "Base_datos",
        "key": 281,
        "rule": "^Base_datos.*",
    },
    {
        "name": "Texto",
        "key": 276,
        "rule": ".*",
    },
]

def getRuta(path):
    rutas = read_column_excel_as_datatable(".\\Generate\\Settings.xlsx", "Rutas",0,1)
    return str(rutas[path][0]) 

# function check if string match to regex
def validate_regex( description, regex):
    # convert string to regex
    regex = re.compile(regex, re.IGNORECASE)

    if ( re.search(regex, str(description)) ):
        return True
    else:
        return False


# function to read excel and return a dataframe
def read_excel(file_path, sheet_name="Settings"):
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    # clean rows with empty values
    df = df.dropna(how="all")
    return df


def read_column_excel_as_datatable(
    file_path, sheet_name="Settings", column_key=0, column_value=2
):
    df = read_excel(file_path, sheet_name)

    # Initialize new datatable "config"
    result = pd.DataFrame()

    for index, row in df.iterrows():
        # add new column to datatable
        result.loc[0, row[column_key]] = row[column_value]

    return result.dropna(how="all")


# function write append to txt file
def write_txt(file_path, text):
    with open(file_path, "a", encoding="utf-8") as f:
        f.write(str(text) + "\n")
        f.close()


# Function return the template
def get_template(name):
    templates = read_column_excel_as_datatable(
        str(getRuta("pathConfigGenerate")), "Templates", 0, 1
    )
    return str(templates[name][0])


def generate_queries_config():
    config_rpa = read_excel(str(getRuta("configRPA")))
    config_generate = read_column_excel_as_datatable(
        str(getRuta("pathConfigGenerate")), "Settings", 0, 2
    )
    path_output = str(getRuta("pathOutputQueries")).replace(
        "9999", str(config_generate["GOC OC"][0])
    )

    template1 = (
        get_template("template1")
        .replace("$1", str(config_rpa["ProcesoID"][0]))
        .replace("{id_negocio}", str(config_generate["ID Negocio"][0]))
        .replace(
            "{name_robot}", str(config_generate["nombre robot"][0]).lower().title()
        )
        .replace(
            "{name_robot_underscore}",
            str(config_generate["nombre robot"][0]).upper().replace(" ", "_"),
        )
    )

    create_directory_production_if_not_exist()
    write_txt(path_output, template1)

    # for to dataframe and get the values
    for index, row in config_rpa.iterrows():
        for match in matchs:
            # if regex match

            if validate_regex(
                 row["Descripcion"], match["rule"]
            ):
                newRow = get_template("template3")
                newRow = newRow.replace("$1", str(match["key"]))
                newRow = newRow.replace("$2", str(row["CodigoUIpath"]))
                newRow = newRow.replace("$3", str(row["Descripcion"]))
                newRow = newRow.replace("$4", str(row["Valor"]).replace("'", "''"))
                write_txt(path_output, newRow)
                break


# Genera el documento PR, REV, PO
def generate_document_PR_REV_PO():
    today = datetime.now().strftime("%d/%m/%Y")
    settings = read_column_excel_as_datatable(
        str(getRuta("pathConfigGenerate")), "Settings", 0, 2
    )

    payload = {
        "name_robot": settings["nombre robot"][0].lower().title(),
        "Desarrollador": settings["Desarrollador"][0],
        "Today": today,
        "GOC OT": settings["GOC OT"][0],
        "GOC OC": settings["GOC OC"][0],
        "Full GOC": settings["GOC OC"][0],
        "Custodio": settings["Custodio"][0],
        "Sistema o Aplicativo": settings["Sistema o Aplicativo"][0],
    }
    create_directory_production_if_not_exist()

    # Prepare document PR
    pr = str(getRuta("pathTemplates")) + "PR_GH_GOC-9999.xlsx"
    wb = load_workbook(pr, read_only=False, keep_vba=True)

    ws = wb["Cabecera"]
    ws["E10"] = payload["Desarrollador"]
    ws["B10"] = payload["Today"]

    ws = wb["RESUMEN"]
    ws["D5"] = payload["Full GOC"]
    ws["D7"] = payload["Custodio"]

    ws = wb["Base de Datos"]
    ws["D123"] = "GOC-" + payload["GOC OC"]
    ws["E123"] = settings["link script"][0]
    ws["F123"] = "SCRIPT_GOC-9999_1.sql".replace("9999", payload["GOC OC"])
    ws["E135"] = (
        "'https://jira.sura.net.pe/confluence/pages/viewpage.action?spaceKey=RPA&title=GOC_"
        + payload["GOC OC"]
    )

    wb.save(
         str(getRuta("pathProduction"))
        + "PR_GH_GOC-9999.xlsm".replace("9999", payload["GOC OC"])
    )

    # Prepare document REV
    rev = str(getRuta("pathTemplates")) + "REV_GH_GOC-9999.xlsm"
    wb = load_workbook(rev, read_only=False, keep_vba=True)

    ws = wb["Historial de Revisiones"]
    ws["G7"] = payload["Desarrollador"]
    ws["G8"] = payload["Custodio"]
    ws["G6"] = payload["Sistema o Aplicativo"]
    ws["D15"] = payload["Today"]
    ws["G4"] = "GOC-" + payload["GOC OT"]
    ws["G5"] = "GOC-" + payload["GOC OC"]

    ws = wb["CHK Técnico-SQL"]
    ws["I19"] = payload["Today"]

    ws = wb["CHK Técnico-RPA"]
    ws["I19"] = payload["Today"]

    wb.save(
         str(getRuta("pathProduction"))
        + "REV_GH_GOC-9999.xlsm".replace("9999", payload["GOC OC"])
    )

    # Prepare document PO
    po = str(getRuta("pathTemplates")) + "PO_GH_GOC-9999.xlsm"
    wb = load_workbook(po, read_only=False, keep_vba=True)

    ws = wb["Historial de Revisiones"]
    ws["D9"] = payload["Today"]

    ws = wb["General"]
    ws["C8"] = "GOC-" + payload["GOC OT"]
    ws["C9"] = "GOC-" + payload["GOC OC"]

    wb.save(
         str(getRuta("pathProduction"))
        + "PO_GH_GOC-9999.xlsm".replace("9999", payload["GOC OC"])
    )


def create_directory_production_if_not_exist():
    # create directory if not exist
    if not os.path.exists( str(getRuta("pathProduction"))):
        os.makedirs( str(getRuta("pathProduction")))

def create_document_LR():
    create_directory_production_if_not_exist()
    today = datetime.now().strftime("%d/%m/%Y")
    settings = read_column_excel_as_datatable(
        str(getRuta("pathConfigGenerate")), "Settings", 0, 2
    )

    payload = {
        "name_robot": settings["nombre robot"][0].lower().title(),
        "Desarrollador": settings["Desarrollador"][0],
        "Today": today,
        "GOC OT": settings["GOC OT"][0],
        "GOC OC": settings["GOC OC"][0],
        "Full GOC": settings["GOC OC"][0],
        "Custodio": settings["Custodio"][0],
        "Sistema o Aplicativo": settings["Sistema o Aplicativo"][0],
    }

    # Prepare document LR
    lr = str(getRuta("pathTemplates")) + "LR_GH_GOC-9999.xlsm"
    wb = load_workbook(lr, read_only=False, keep_vba=True)

    ws = wb["Historial de Revisiones"]
    ws["D8"] = payload["Today"]

    ws = wb["Requerimientos de usuario"]
    ws["E9"] = payload["Custodio"]


    ws = wb["Requerimientos del sistema"]
    ws["F10"] = payload["Full GOC"]


    wb.save(
         str(getRuta("pathProduction"))
        + "LR_GH_GOC-9999.xlsm".replace("9999", payload["GOC OC"])
    )
    wb.close


def captureRequerimentsToSource():
    workflows = find_workflows(str(getRuta("pathRPACodigo"))+"Main.xaml")
    sorted_array = sorted(workflows, key=lambda x: ('InitAllSettings' in x, 'InitAllApplications' in x, 'Process' in x))
    sorted_array.reverse()

    for workflow in sorted_array:
        print(workflow)
        

def find_workflows(fullfile):
    regex_workflows = r".+FileName=\"(.*\.xaml)\""
    used_workflows = []

    # Open file as file object and read to string
    file = open(f"{fullfile}", "r")

    # Read file object to string
    text = file.read()

    # Close file object
    file.close()

    # Regex pattern
    pattern = re.compile(
        regex_workflows,
        re.MULTILINE,
    )

    workflows = pattern.finditer(text)

    for file in workflows:
        used_workflows.append(str(getRuta("pathRPACodigo")) + file.group(1))
        find_workflows(str(getRuta("pathRPACodigo"))+file.group(1))
    
    return used_workflows


def create_document_CP():
    create_directory_production_if_not_exist()
    today = datetime.now().strftime("%d/%m/%Y")
    settings = read_column_excel_as_datatable(
        str(getRuta("pathConfigGenerate")), "Settings", 0, 2
    )

    payload = {
        "name_robot": settings["nombre robot"][0].lower().title(),
        "Desarrollador": settings["Desarrollador"][0],
        "Today": today,
        "GOC OT": settings["GOC OT"][0],
        "GOC OC": settings["GOC OC"][0],
        "Full GOC": settings["GOC OC"][0],
        "Custodio": settings["Custodio"][0],
        "Sistema o Aplicativo": settings["Sistema o Aplicativo"][0],
    }

    # Prepare document CP
    cp = str(getRuta("pathTemplates")) +  "CP_GH_GOC-9999.xlsm"
    wb = load_workbook(cp, read_only=False, keep_vba=True)

    ws = wb["Historial de Revisiones"]
    ws["C6"] = payload["GOC OT"]
    ws["C7"] = payload["GOC OC"]
    ws["C10"] = payload["Today"]
    ws["C17"] = payload["Today"]

    ws = wb["Resumen Escenarios"]
    ws["C3"] = payload["name_robot"]

    casos_prueba = read_excel(str(getRuta("pathConfigGenerate")),"Casos de uso")
    numero_fila = 14
    total = 0

    for index, row in casos_prueba.iterrows():
        ws["B"+str( numero_fila + index)] = str(index)
        ws["C"+str( numero_fila + index)] = "ES00" + str(index)
        ws["D"+str( numero_fila + index)] = str(row["Caso de prueba"])
        ws["E"+str( numero_fila + index)] = str(row["Resultados Esperado"])
        total = index
    
    ws["C8"] = str(total)

    wb.save(
         str(getRuta("pathProduction"))
        + "CP_GH_GOC-9999.xlsm".replace("9999", payload["GOC OC"])
    )

    wb.close


def generate_docx_production():
    path_template = str(getRuta("pathTemplates")) + "Production Document.docx"
    path_config_rpa = str(getRuta("configRPA"))

    # Prepare template an data
    doc = DocxTemplate(path_template)
    config_generate = read_column_excel_as_datatable(
        str(getRuta("pathConfigGenerate")), "Settings", 0, 2
    )

    # for to dataframe and get the values
    for index, row in config_generate.iterrows():

        links = re.findall(r"(\d{3,6})", row["link confluence"])[0]
        if links:
            goc = links
        else:
            goc = 9999

        # Prepare payload for template
        payload = {
            "name_robot": row["nombre robot"].lower().title(),
            "name_robot_under_score": row["nombre robot"].upper().replace(" ", "_"),
            "version_robot": row["version nupkg"],
            "name_empresa": row["empresa"],
            "link_goc": row["link bitbucket"],
            "link_confluence": row["link confluence"],
            "link_nupkg": row["link nupkg"],
            "link_config": row["link config"],
            "cron_expression": row["cron"],
            "host_prod": row["Host prod"],
            "NAME_ROBOT": row["nombre robot"].upper(),
            "NAME_ROBOT_UNDER_SCORE": row["nombre robot"].upper().replace(" ", "_"),
            "NAME_EMPRESA": row["empresa"].upper(),
            "goc": goc,
            "assets": [],
            "ficheros": [],
        }

        # List all directories recursively, ingnoring hidden directories and save  in payload["ficheros"]
        path_project = os.getcwd()
        for root, dirs, files in os.walk(os.getcwd()):
            for dir in dirs:
                # is string not contains .git
                if ".git" not in str(os.path.join(root, dir)) and "Fuente" not in str(
                    os.path.join(root, dir)
                ):
                    folder = str(os.path.join(root, dir)).replace(
                        path_project + "\\", ""
                    )

                    payload["ficheros"].append({"name": folder})

        # Prepare name with underscore
        robot_score = row["nombre robot"].upper().replace(" ", "_")
        path_output = "RPA_{name_robot} - Production Document.docx".replace(
            "{name_robot}", robot_score
        )

        # Read excel Config of RPA
        rpa_config = read_excel(path_config_rpa)

        # Get assets to config.xlsx
        for index, row in rpa_config.iterrows():
            if validate_regex(
               row["Descripcion"], matchs[5]["rule"]
            ):
                payload["assets"].append({"name": row["Valor"]})

        # Generate documents
        jinja_env = jinja2.Environment(autoescape=True)
        doc.render(payload, jinja_env)

        create_directory_production_if_not_exist()
        doc.save( str(getRuta("pathProduction")) + path_output)


menu_options = {
    1: "Generar script para subir el Config a la base de datos",
    2: "Generar documento word para el pase a producción",
    3: "Generar documentos PR, REV y PO",
    4: "create_document_CP",
    99: "Ejecutar todo",
    0: "Salir",
}


def print_menu():
    for key in menu_options.keys():
        print(key, ": ", menu_options[key])


def message_success():
    print("Proceso terminado con éxito \n")


# Main

if __name__ == "__main__":
    while True:
        print_menu()
        option = ""
        try:
            option = int(input("Eliga una opción: "))
        except:
            print("Error, por favor ingrese un numero...")
        # Check what choice was entered and act accordingly
        if option == 1:
            generate_queries_config()
            message_success()
        elif option == 2:
            generate_docx_production()
            message_success()
        elif option == 3:
            generate_document_PR_REV_PO()
            message_success()
        elif option == 4:
            create_document_CP()
        elif option == 5:
            captureRequerimentsToSource()

        elif option == 99:
            generate_docx_production()
            generate_queries_config()
            generate_document_PR_REV_PO()
            exit()
        elif option == 0:
            print("Saliendo")
            exit()

        else:
            print("Opción invalida")
